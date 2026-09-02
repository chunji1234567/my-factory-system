"""从 xlsx 导入合作方到新系统（2026-08-21）。

用法（在服务器上）：
    # 1. 把 合作方名单.xlsx 上传到服务器（放到 /srv/my-factory-system/my-factory-system/ 下）
    # 2. 把本脚本也放到同目录
    # 3. 跑：
    cd /srv/my-factory-system/my-factory-system
    backend/venv/bin/python manage.py shell < scripts/import_partners.py

规则：
    - Excel 第 1 列 = 姓名，第 2 列 = 公司名称，第 3 列 = 类型（客户 / 供应商）
    - 名字优先级：有公司名用公司名，没公司名用姓名（去 strip）
    - 类型映射：客户 → CUSTOMER，供应商 → SUPPLIER
    - 幂等：按最终名字去重，已存在的合作方跳过（不覆盖）
    - 有空行/无类型的行 → 打印警告并跳过

需要修改（如果 xlsx 不在 my-factory-system/ 根目录）：
    XLSX_PATH 改成实际路径
"""
from pathlib import Path
from openpyxl import load_workbook

from core.models import Partner


# ============================================================================
# 修改这里：xlsx 文件路径
# ============================================================================
XLSX_PATH = '合作方名单.xlsx'   # 假设放在项目根目录


TYPE_MAP = {
    '客户': 'CUSTOMER',
    '供应商': 'SUPPLIER',
    '两者': 'BOTH',
    '兼有': 'BOTH',
}


def main():
    if not Path(XLSX_PATH).exists():
        print(f'❌ 找不到文件：{XLSX_PATH}')
        print(f'   当前目录：{Path.cwd()}')
        return

    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb.active   # 第一个 sheet

    created = 0
    skipped = 0
    warned = 0
    dup_names = set()

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        # 跳过表头
        if i == 1:
            continue
        # 空行
        if not any(row):
            continue

        name = (row[0] or '').strip() if row[0] else ''
        company = (row[1] or '').strip() if len(row) > 1 and row[1] else ''
        raw_type = (row[2] or '').strip() if len(row) > 2 and row[2] else ''

        # 最终名字：优先公司，没公司用姓名
        final_name = company or name
        if not final_name:
            print(f'  ⚠️  行 {i}：姓名和公司都是空，跳过')
            warned += 1
            continue

        # 类型映射
        partner_type = TYPE_MAP.get(raw_type)
        if not partner_type:
            print(f'  ⚠️  行 {i}（{final_name}）：类型 "{raw_type}" 无法识别，跳过')
            warned += 1
            continue

        # 同一批 xlsx 里重名检查
        if final_name in dup_names:
            print(f'  ⚠️  行 {i}（{final_name}）：本次导入内有重名，跳过第二次')
            warned += 1
            continue
        dup_names.add(final_name)

        # 幂等：数据库已有同名 → 跳过（不覆盖）
        obj, is_new = Partner.objects.get_or_create(
            name=final_name,
            defaults={'partner_type': partner_type},
        )
        if is_new:
            print(f'  ✓ 行 {i}：新建 [{partner_type}] {final_name}')
            created += 1
        else:
            print(f'  · 行 {i}：已存在 {final_name}（type={obj.partner_type}），跳过')
            skipped += 1

    print()
    print(f'完成。新建 {created} 个，已存在跳过 {skipped} 个，警告 {warned} 个。')


main()
