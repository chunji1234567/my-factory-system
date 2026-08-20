"""经营分析 API（2026-06-20 新增，2026-08-21 扩展）。

manager only。聚合 endpoint：
  * /api/business/analytics/trend/
    销售 vs 采购金额按时间分组的时序数据。
  * /api/business/analytics/material-consumption/
    "实际消耗"口径——从 PRODUCE_CONSUME StockAdjustment 聚合，
    反映车间已经扣过的料。支持 category 过滤。
  * /api/business/analytics/material-consumption/export/
    物料消耗表 xlsx 导出（同上口径，含 category 过滤）。
  * /api/business/analytics/production-throughput/
    排产吞吐量按时间分组的时序数据（sum(ProductionRecord.quantity)）。
  * /api/business/analytics/plan-sales/       [2026-08-21]
    "订单需求"侧的方案销售——每个 PCB 方案在期间的销售数量 + 金额，
    数据源 SalesOrderItem（按 order.created_at 过滤）。
  * /api/business/analytics/material-demand/  [2026-08-21]
    "订单需求"侧的物料展开——外壳/线材直接按 sales_item.quantity，
    方案按 sales_item.quantity × PcbPlanMaterial.quantity_per_unit 展开到原料。
    与 material-consumption 是互补视角：一个是"客户下单要多少料"，
    一个是"车间实际扣了多少料"。差值 = 待生产 + 排产滞后。

公共查询参数：
  from        起始日期 YYYY-MM-DD（必填）
  to          结束日期 YYYY-MM-DD，含当天（必填）
  group_by    day / week / month（趋势 API 用）
  category    (可选) 分类 id——material-consumption / material-demand 支持

时区注意：所有 created_at / executed_at 都是 UTC 存储，
但对老板来说"6 月销售"就是 UTC+8 6 月 1 日到 30 日。这里
直接按 date-part 分组，够用。真正精细的时区处理留给日后。

详见 docs/PRD.md §9.4 changelog 2026-06-20（经营分析）+ 2026-08-21（订单需求视图）。
"""
from __future__ import annotations

import io
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal

from django.db.models import Sum, F, DecimalField, Value, ExpressionWrapper, Q
from django.db.models.functions import Coalesce, TruncDay, TruncWeek, TruncMonth
from django.http import HttpResponse
from django.utils.dateparse import parse_date
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from business.models import (
    SalesOrder,
    SalesOrderItem,
    PurchaseOrder,
    PurchaseOrderItem,
    ProductionRecord,
    StockAdjustment,
)
from core.models import Product, Category, PcbPlan, PcbPlanMaterial
from .permissions import IsManager


# ---------------------------------------------------------------------------
# 公共工具
# ---------------------------------------------------------------------------
def _parse_range(request):
    """从 query params 解析 from/to 日期，返回 (from_date, to_date, error_response)。

    from/to 必填。error_response 非 None 时调用方直接返回它。
    """
    from_str = request.query_params.get('from')
    to_str = request.query_params.get('to')
    if not from_str or not to_str:
        return None, None, Response(
            {'detail': '需要提供 from 和 to 查询参数（YYYY-MM-DD 格式）'},
            status=400,
        )
    from_date = parse_date(from_str)
    to_date = parse_date(to_str)
    if not from_date or not to_date:
        return None, None, Response(
            {'detail': '无效的日期格式，请用 YYYY-MM-DD'},
            status=400,
        )
    if from_date > to_date:
        return None, None, Response(
            {'detail': 'from 不能晚于 to'},
            status=400,
        )
    return from_date, to_date, None


def _get_trunc(group_by: str):
    """把 group_by 字符串映射到 Django Trunc 函数 + 前端 label 格式。"""
    if group_by == 'week':
        return TruncWeek, '%Y-W%V'  # ISO 周
    if group_by == 'month':
        return TruncMonth, '%Y-%m'
    # 默认按天
    return TruncDay, '%Y-%m-%d'


# ---------------------------------------------------------------------------
# 1. 销售 vs 采购趋势
# ---------------------------------------------------------------------------
class SalesPurchaseTrendView(APIView):
    permission_classes = [IsAuthenticated, IsManager]

    def get(self, request):
        from_date, to_date, err = _parse_range(request)
        if err:
            return err
        group_by = request.query_params.get('group_by', 'day')
        Trunc, date_fmt = _get_trunc(group_by)

        # 聚合销售
        sales_qs = (
            SalesOrder.objects
            .filter(created_at__date__gte=from_date, created_at__date__lte=to_date)
            .annotate(period=Trunc('created_at'))
            .values('period')
            .annotate(total=Sum('total_amount'))
            .order_by('period')
        )
        # 聚合采购
        purchase_qs = (
            PurchaseOrder.objects
            .filter(created_at__date__gte=from_date, created_at__date__lte=to_date)
            .annotate(period=Trunc('created_at'))
            .values('period')
            .annotate(total=Sum('total_amount'))
            .order_by('period')
        )

        # 合并到同一时间轴：先拿全部 period 键的并集，然后按顺序输出
        sales_map = {r['period'].date(): r['total'] or Decimal('0') for r in sales_qs}
        purchase_map = {r['period'].date(): r['total'] or Decimal('0') for r in purchase_qs}
        all_periods = sorted(set(sales_map.keys()) | set(purchase_map.keys()))

        return Response({
            'from': from_date.isoformat(),
            'to': to_date.isoformat(),
            'group_by': group_by,
            'series': [
                {
                    'period': p.strftime(date_fmt),
                    'period_date': p.isoformat(),
                    'sales': float(sales_map.get(p, Decimal('0'))),
                    'purchase': float(purchase_map.get(p, Decimal('0'))),
                }
                for p in all_periods
            ],
            'totals': {
                'sales': float(sum(sales_map.values())),
                'purchase': float(sum(purchase_map.values())),
            },
        })


# ---------------------------------------------------------------------------
# 2. 物料消耗
# ---------------------------------------------------------------------------
def _material_consumption_rows(from_date, to_date, category_id=None):
    """核心聚合：返回 [{material, quantity, avg_price, total_value, daily_avg}, ...]

    抽出函数便于 xlsx 导出复用。

    :param category_id: 可选分类过滤，只保留属于该 Category 的物料。
        （直接按 category id 过 Product 表；不做跨父分类展开。）
    """
    days = max(1, (to_date - from_date).days + 1)

    # ① 消耗数量：从 PRODUCE_CONSUME StockAdjustment 聚合
    consumed_qs = (
        StockAdjustment.objects
        .filter(
            adjustment_type='PRODUCE_CONSUME',
            created_at__date__gte=from_date,
            created_at__date__lte=to_date,
        )
    )
    if category_id:
        # PRODUCE_CONSUME 为出库，quantity 为负，先按 product 聚合再过滤 category
        consumed_qs = consumed_qs.filter(product__category_id=category_id)
    consumed = (
        consumed_qs
        .values('product')
        .annotate(quantity=Sum('quantity'))
    )
    consumed_map = {r['product']: r['quantity'] or Decimal('0') for r in consumed}

    if not consumed_map:
        return []

    # ② 加权平均采购价：SUM(price × quantity) / SUM(quantity) from PurchaseOrderItem
    # 期间内的进货加权。若期间无进货则回退到全时期加权。
    price_agg = (
        PurchaseOrderItem.objects
        .filter(
            product_id__in=consumed_map.keys(),
            order__created_at__date__gte=from_date,
            order__created_at__date__lte=to_date,
        )
        .values('product')
        .annotate(
            total_value=Sum(ExpressionWrapper(
                F('price') * F('quantity'),
                output_field=DecimalField(max_digits=20, decimal_places=4),
            )),
            total_qty=Sum('quantity'),
        )
    )
    price_map = {}
    for r in price_agg:
        if r['total_qty'] and r['total_qty'] > 0:
            price_map[r['product']] = r['total_value'] / r['total_qty']

    # 期间内无进货 → 全时期回退
    missing_price_products = [
        pid for pid in consumed_map.keys() if pid not in price_map
    ]
    if missing_price_products:
        fallback_agg = (
            PurchaseOrderItem.objects
            .filter(product_id__in=missing_price_products)
            .values('product')
            .annotate(
                total_value=Sum(ExpressionWrapper(
                    F('price') * F('quantity'),
                    output_field=DecimalField(max_digits=20, decimal_places=4),
                )),
                total_qty=Sum('quantity'),
            )
        )
        for r in fallback_agg:
            if r['total_qty'] and r['total_qty'] > 0:
                price_map[r['product']] = r['total_value'] / r['total_qty']

    # ③ 拉物料元信息
    products = Product.objects.filter(
        id__in=consumed_map.keys()
    ).select_related('category').only(
        'id', 'model_name', 'internal_code', 'category__id', 'category__name',
    )
    product_map = {p.id: p for p in products}

    # ④ 组装行
    rows = []
    for pid, qty in consumed_map.items():
        p = product_map.get(pid)
        if not p:
            continue
        avg_price = price_map.get(pid, Decimal('0'))
        total_value = qty * avg_price
        daily_avg = qty / Decimal(days)
        rows.append({
            'product_id': pid,
            'model_name': p.model_name,
            'internal_code': p.internal_code,
            'category': getattr(p.category, 'name', '') or '',
            'category_id': getattr(p.category, 'id', None),
            'quantity': qty,
            'avg_price': avg_price,
            'total_value': total_value,
            'daily_avg': daily_avg,
        })

    # 默认按消耗数量降序
    rows.sort(key=lambda r: r['quantity'], reverse=True)
    return rows


def _parse_category(request):
    """解析 ?category=<id>，非法值返回 None（视为不过滤，不报错）。"""
    raw = request.query_params.get('category')
    if not raw:
        return None
    try:
        return int(raw)
    except (TypeError, ValueError):
        return None


def _category_options():
    """所有分类扁平列表——给前端下拉用。按 category_type 分组显示时前端自行处理。"""
    return [
        {
            'id': c.id,
            'name': c.name,
            'category_type': c.category_type,
            'category_type_display': c.get_category_type_display(),
        }
        for c in Category.objects.all().order_by('category_type', 'name')
    ]


class MaterialConsumptionView(APIView):
    permission_classes = [IsAuthenticated, IsManager]

    def get(self, request):
        from_date, to_date, err = _parse_range(request)
        if err:
            return err
        category_id = _parse_category(request)
        rows = _material_consumption_rows(from_date, to_date, category_id=category_id)
        return Response({
            'from': from_date.isoformat(),
            'to': to_date.isoformat(),
            'category': category_id,
            'categories': _category_options(),
            'rows': [
                {
                    'product_id': r['product_id'],
                    'model_name': r['model_name'],
                    'internal_code': r['internal_code'],
                    'category': r['category'],
                    'category_id': r.get('category_id'),
                    'quantity': float(r['quantity']),
                    'avg_price': float(r['avg_price']),
                    'total_value': float(r['total_value']),
                    'daily_avg': float(r['daily_avg']),
                }
                for r in rows
            ],
            'summary': {
                'total_value': float(sum(r['total_value'] for r in rows)),
                'total_quantity': float(sum(r['quantity'] for r in rows)),
                'material_count': len(rows),
            },
        })


class MaterialConsumptionExportView(APIView):
    """物料消耗 xlsx 导出（支持 category 过滤，与列表口径一致）。"""
    permission_classes = [IsAuthenticated, IsManager]

    def get(self, request):
        from_date, to_date, err = _parse_range(request)
        if err:
            return err
        category_id = _parse_category(request)
        rows = _material_consumption_rows(from_date, to_date, category_id=category_id)

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = '物料消耗'

        headers = ['物料名', '内部编号', '分类', '消耗数量', '加权单价', '总价值', '日均消耗']
        n_cols = len(headers)

        # 第 1 行：标题
        title = f'物料消耗汇总  {from_date}  至  {to_date}'
        ws.cell(row=1, column=1, value=title)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(name='Microsoft YaHei', size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill('solid', fgColor='F1F5F9')
        ws.row_dimensions[1].height = 28

        # 第 2 行：表头
        header_font = Font(name='Microsoft YaHei', bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='475569')
        header_align = Alignment(horizontal='center', vertical='center')
        for col_idx, label in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        ws.row_dimensions[2].height = 22

        body_font = Font(name='Microsoft YaHei', size=11)
        left = Alignment(horizontal='left', vertical='center')
        center = Alignment(horizontal='center', vertical='center')
        right = Alignment(horizontal='right', vertical='center')
        money_fmt = '#,##0.00'
        qty_fmt = '#,##0.##'

        row = 3
        for r in rows:
            ws.cell(row=row, column=1, value=r['model_name']).alignment = left
            ws.cell(row=row, column=2, value=r['internal_code']).alignment = center
            ws.cell(row=row, column=3, value=r['category']).alignment = center
            c4 = ws.cell(row=row, column=4, value=float(r['quantity']))
            c4.number_format = qty_fmt
            c4.alignment = right
            c5 = ws.cell(row=row, column=5, value=float(r['avg_price']))
            c5.number_format = money_fmt
            c5.alignment = right
            c6 = ws.cell(row=row, column=6, value=float(r['total_value']))
            c6.number_format = money_fmt
            c6.alignment = right
            c7 = ws.cell(row=row, column=7, value=float(r['daily_avg']))
            c7.number_format = qty_fmt
            c7.alignment = right
            for c in range(1, n_cols + 1):
                ws.cell(row=row, column=c).font = body_font
            row += 1

        # 自适应列宽
        for col_idx in range(1, n_cols + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[col_letter]:
                if cell.value is None:
                    continue
                s = str(cell.value)
                width = sum(1.7 if ord(ch) > 127 else 1.0 for ch in s)
                if width > max_len:
                    max_len = width
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)

        ws.freeze_panes = 'A3'

        buffer = io.BytesIO()
        wb.save(buffer)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        range_str = f'{from_date.strftime("%Y%m%d")}_{to_date.strftime("%Y%m%d")}'
        filename_ascii = f'material_consumption_{range_str}.xlsx'
        filename_utf8 = f'物料消耗_{range_str}.xlsx'
        from urllib.parse import quote
        response['Content-Disposition'] = (
            f'attachment; filename="{filename_ascii}"; '
            f"filename*=UTF-8''{quote(filename_utf8)}"
        )
        return response


# ---------------------------------------------------------------------------
# 3. 排产吞吐量趋势
# ---------------------------------------------------------------------------
class ProductionThroughputView(APIView):
    permission_classes = [IsAuthenticated, IsManager]

    def get(self, request):
        from_date, to_date, err = _parse_range(request)
        if err:
            return err
        group_by = request.query_params.get('group_by', 'day')
        Trunc, date_fmt = _get_trunc(group_by)

        qs = (
            ProductionRecord.objects
            .filter(executed_at__date__gte=from_date, executed_at__date__lte=to_date)
            .annotate(period=Trunc('executed_at'))
            .values('period')
            .annotate(total=Sum('quantity'))
            .order_by('period')
        )
        series = [
            {
                'period': r['period'].strftime(date_fmt),
                'period_date': r['period'].date().isoformat(),
                'quantity': float(r['total'] or 0),
            }
            for r in qs
        ]
        return Response({
            'from': from_date.isoformat(),
            'to': to_date.isoformat(),
            'group_by': group_by,
            'series': series,
            'total': sum(s['quantity'] for s in series),
        })


# ---------------------------------------------------------------------------
# 4. 方案销售（订单需求侧，2026-08-21）
# ---------------------------------------------------------------------------
class PlanSalesView(APIView):
    """每个 PCB 方案在期间的销售数量 + 销售金额。

    数据源：SalesOrderItem，按 order.created_at 过滤。含 pcb_plan 的行才计入。
    金额 = SUM(price × quantity)。
    """
    permission_classes = [IsAuthenticated, IsManager]

    def get(self, request):
        from_date, to_date, err = _parse_range(request)
        if err:
            return err

        # 走 Python 端聚合——避免 Sum(ExpressionWrapper) 在某些 Django 版本上
        # 分组聚合触发的 FieldError。数据量在方案 × 期间订单数级别，很小。
        items = (
            SalesOrderItem.objects
            .filter(
                pcb_plan__isnull=False,
                order__created_at__date__gte=from_date,
                order__created_at__date__lte=to_date,
            )
            .values_list('pcb_plan_id', 'quantity', 'price')
        )

        qty_map: dict[int, Decimal] = defaultdict(lambda: Decimal('0'))
        rev_map: dict[int, Decimal] = defaultdict(lambda: Decimal('0'))
        for plan_id, qty, price in items:
            qty = qty or Decimal('0')
            price = price or Decimal('0')
            qty_map[plan_id] += qty
            rev_map[plan_id] += qty * price

        plan_map = {p.id: p for p in PcbPlan.objects.filter(id__in=qty_map.keys())}

        rows = []
        for plan_id, qty in qty_map.items():
            plan = plan_map.get(plan_id)
            if not plan:
                continue
            rev = rev_map[plan_id]
            avg_price = (rev / qty) if qty > 0 else Decimal('0')
            rows.append({
                'plan_id': plan.id,
                'plan_name': plan.name,
                'plan_code': plan.code,
                'is_active': plan.is_active,
                'quantity': float(qty),
                'revenue': float(rev),
                'avg_price': float(avg_price),
            })
        rows.sort(key=lambda r: r['quantity'], reverse=True)

        return Response({
            'from': from_date.isoformat(),
            'to': to_date.isoformat(),
            'rows': rows,
            'summary': {
                'plan_count': len(rows),
                'total_quantity': float(sum(Decimal(str(r['quantity'])) for r in rows)),
                'total_revenue': float(sum(Decimal(str(r['revenue'])) for r in rows)),
            },
        })


# ---------------------------------------------------------------------------
# 5. 订单展开的物料需求（2026-08-21）
# ---------------------------------------------------------------------------
def _material_demand_rows(from_date, to_date, category_id=None):
    """从 SalesOrderItem 展开外壳 / 线材 / 方案 → 物料需求量。

    展开规则：
      - product（外壳，SELF_MADE） → 1:1，qty = SUM(item.quantity)
      - cable（CABLE） → 1:1，qty = SUM(item.quantity)
      - pcb_plan → 该方案下每种 material，qty += item.quantity × material.quantity_per_unit

    单价用采购加权平均，逻辑与实际消耗一致：期间无进货则全时期回退。
    """
    items = (
        SalesOrderItem.objects
        .filter(
            order__created_at__date__gte=from_date,
            order__created_at__date__lte=to_date,
        )
        .values('id', 'quantity', 'product_id', 'cable_id', 'pcb_plan_id')
    )
    items = list(items)
    if not items:
        return []

    # 累计每种 product 的需求量
    demand_map: dict[int, Decimal] = defaultdict(lambda: Decimal('0'))
    # 用于统计每种物料被哪些槽位需求（便于前端展示"来自：外壳/线材/方案"）
    source_map: dict[int, set[str]] = defaultdict(set)

    # 收集用到的 pcb_plan → 提前展开物料
    plan_ids = {it['pcb_plan_id'] for it in items if it['pcb_plan_id']}
    plan_materials: dict[int, list[tuple[int, Decimal]]] = defaultdict(list)
    if plan_ids:
        for pm in PcbPlanMaterial.objects.filter(plan_id__in=plan_ids).values(
            'plan_id', 'material_id', 'quantity_per_unit',
        ):
            plan_materials[pm['plan_id']].append(
                (pm['material_id'], pm['quantity_per_unit'] or Decimal('0'))
            )

    for it in items:
        qty = it['quantity'] or Decimal('0')
        if qty <= 0:
            continue
        if it['product_id']:
            demand_map[it['product_id']] += qty
            source_map[it['product_id']].add('shell')
        if it['cable_id']:
            demand_map[it['cable_id']] += qty
            source_map[it['cable_id']].add('cable')
        pid = it['pcb_plan_id']
        if pid and pid in plan_materials:
            for mat_id, per_unit in plan_materials[pid]:
                demand_map[mat_id] += qty * per_unit
                source_map[mat_id].add('plan')

    if not demand_map:
        return []

    # 加权采购价（同 _material_consumption_rows 口径）
    price_agg = (
        PurchaseOrderItem.objects
        .filter(
            product_id__in=demand_map.keys(),
            order__created_at__date__gte=from_date,
            order__created_at__date__lte=to_date,
        )
        .values('product')
        .annotate(
            total_value=Sum(ExpressionWrapper(
                F('price') * F('quantity'),
                output_field=DecimalField(max_digits=20, decimal_places=4),
            )),
            total_qty=Sum('quantity'),
        )
    )
    price_map = {}
    for r in price_agg:
        if r['total_qty'] and r['total_qty'] > 0:
            price_map[r['product']] = r['total_value'] / r['total_qty']

    missing_price = [pid for pid in demand_map.keys() if pid not in price_map]
    if missing_price:
        fallback = (
            PurchaseOrderItem.objects
            .filter(product_id__in=missing_price)
            .values('product')
            .annotate(
                total_value=Sum(ExpressionWrapper(
                    F('price') * F('quantity'),
                    output_field=DecimalField(max_digits=20, decimal_places=4),
                )),
                total_qty=Sum('quantity'),
            )
        )
        for r in fallback:
            if r['total_qty'] and r['total_qty'] > 0:
                price_map[r['product']] = r['total_value'] / r['total_qty']

    # 元信息 + 可选 category 过滤
    prod_qs = Product.objects.filter(id__in=demand_map.keys()).select_related('category').only(
        'id', 'model_name', 'internal_code', 'category__id', 'category__name',
    )
    if category_id:
        prod_qs = prod_qs.filter(category_id=category_id)
    product_map = {p.id: p for p in prod_qs}

    rows = []
    for pid, qty in demand_map.items():
        p = product_map.get(pid)
        if not p:
            continue
        avg_price = price_map.get(pid, Decimal('0'))
        total_value = qty * avg_price
        rows.append({
            'product_id': pid,
            'model_name': p.model_name,
            'internal_code': p.internal_code,
            'category': getattr(p.category, 'name', '') or '',
            'category_id': getattr(p.category, 'id', None),
            'sources': sorted(source_map[pid]),  # ['shell'|'cable'|'plan', ...]
            'quantity': qty,
            'avg_price': avg_price,
            'total_value': total_value,
        })
    rows.sort(key=lambda r: r['quantity'], reverse=True)
    return rows


class MaterialDemandView(APIView):
    """订单展开的物料需求（与 material-consumption 互补，两者差 ≈ 待生产库存缺口）。"""
    permission_classes = [IsAuthenticated, IsManager]

    def get(self, request):
        from_date, to_date, err = _parse_range(request)
        if err:
            return err
        category_id = _parse_category(request)
        rows = _material_demand_rows(from_date, to_date, category_id=category_id)
        return Response({
            'from': from_date.isoformat(),
            'to': to_date.isoformat(),
            'category': category_id,
            'categories': _category_options(),
            'rows': [
                {
                    'product_id': r['product_id'],
                    'model_name': r['model_name'],
                    'internal_code': r['internal_code'],
                    'category': r['category'],
                    'category_id': r['category_id'],
                    'sources': r['sources'],
                    'quantity': float(r['quantity']),
                    'avg_price': float(r['avg_price']),
                    'total_value': float(r['total_value']),
                }
                for r in rows
            ],
            'summary': {
                'material_count': len(rows),
                'total_quantity': float(sum(r['quantity'] for r in rows)),
                'total_value': float(sum(r['total_value'] for r in rows)),
            },
        })
